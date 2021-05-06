import openpyxl
file_location="enviro.xlsx"
workbook=openpyxl.load_workbook(file_location)
sheet_Users=workbook.get_sheet_by_name('users')
sheet_Rooms=workbook.get_sheet_by_name('rooms')

from tkinter import *
from tkinter import messagebox

root = Tk()
root.title("login")
root.geometry("250x150")
root.resizable(width=False, height=False)

def addUser(username,password):
    #   add new users in db (excel)
    # maxRows_Users = sheet_Users.max_row
    # append to the excel file

    sheet_Users.append([username,password])
    workbook.save(file_location)

def login():
    # process the login info entered onscreen
    strUserName = entUserName.get()
    strUserPassword = entPassword.get()

    lblResult = Label(root, text=" ")

    lgcFound = False

    if len(strUserName) == 0:
        lblResult.grid_forget()
        lblResult = Label(root, text="username cannot be blank")
    else:
        # check if user exists. look in xl file in root
        # if the db is empty, ask if add this user 
        # if user isn't found, ask if add this user
        # if rows are > 1 then there are users
        maxRows_Users = sheet_Users.max_row
        if maxRows_Users > 1:
            for i in range(2, sheet_Users.max_row+1):
                strUserInCell = sheet_Users.cell(i,1).value     
                if strUserInCell==strUserName:
                    lgcFound = True
                    strPasswordInCell = sheet_Users.cell(i,2).value
                    break

            if lgcFound:
                if str(strPasswordInCell) == str(strUserPassword):
                    lblResult = Label(root, text="found " + strUserName)
                else:
                    lblResult = Label(root, text="wrong credentials")
            else:
                result = messagebox.askyesno("user not found", "add " + strUserName + " as new user?")  
                if result:
                    addUser(strUserName,strUserPassword)
                    lblResult = Label(root, text="added new user")
                else:
                    lblResult = Label(root, text="wrong credentials")  
        else:
            # there are no users. add current user to the db
            result = messagebox.askyesno("no users in db", "add as new user?")  
            if  result :           
                addUser(strUserName,strUserPassword)
                lblResult = Label(root, text="added new user")
            else:
                lblResult = Label(root, text="there are no users in the db yet")

    lblResult.grid(row=9, column=0)

# create the interface labels
lblWelcome = Label(root, text="welcome to EnviroControl", padx=55)
lblLogin = Label(root, text="login or new user", padx=55)

lblBlankLine = Label(root, text="      ")

lblUserName = Label(root, text="username:")
lblPassword = Label(root, text="password:")
entUserName = Entry(root, width=15)
entPassword = Entry(root, width=15)

# place labels onscreen
lblWelcome.grid(row=0, column=0, columnspan=3)
lblLogin.grid(row=1, column=0, columnspan=3)
lblBlankLine.grid(row=2, column=0)
lblBlankLine.grid(row=3, column=0)

lblUserName.grid(row=4, column=0)
entUserName.grid(row=4, column=1)

lblPassword.grid(row=5, column=0)
entPassword.grid(row=5, column=1)

lblBlankLine.grid(row=6, column=0)
lblBlankLine.grid(row=7, column=0)

btnLogin = Button(root, text="login", padx=50, command=login, bg="red")
btnLogin.grid(row=8, column=1)

root.mainloop()
